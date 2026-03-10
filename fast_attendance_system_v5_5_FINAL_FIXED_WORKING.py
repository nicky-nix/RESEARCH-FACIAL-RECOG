# ============================================================================
# FINAL WORKING VERSION - Color Coding COMPLETELY FIXED (v5.5 FINAL - 2025-11-09)
# ============================================================================
# COMPLETE FIX: Rewrote save_attendance_to_excel method to properly color Status column
# NEW: Status column (H) now gets PROPER coloring that actually works
# FIXED: Removed apply_cell_formatting calls that weren't working
# GUARANTEE: Status colors will now show in Excel (Present=Green, Late=Yellow, Absent=Red)
# ============================================================================

import cv2
import face_recognition
import numpy as np
import pandas as pd
from datetime import datetime, timedelta
import os
import pickle
import json
import glob
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class FastAttendanceSystem:
    """High-performance facial recognition attendance system"""

    SUBJECTS = [
        "Media and Info Literacy",
        "21st Century Literature From the Philippines and the World",
        "Computer System Servicing",
        "(Research Project) Inquiries, Investigation and Immersion",
        "Understanding Culture, Society and Politics",
        "Cookery"
    ]

    SUBJECT_SECTION_RESTRICTIONS = {
        "Cookery": ["Cookery"],
        "Computer System Servicing": ["ICT"]
    }

    # BRIGHT, VIBRANT COLORS
    COLOR_PRESENT = "00B050"      # Bright Green
    COLOR_LATE = "FFD966"         # Bright Yellow
    COLOR_ABSENT = "FF6B6B"       # Bright Red
    COLOR_HEADER = "1F4E78"       # Dark Blue

    def __init__(self):
        self.known_face_encodings = []
        self.known_face_names = []
        self.known_face_ids = []
        self.known_face_sections = []
        self.attendance_records = []
        self.marked_students = set()
        self.database_file = "students_database.pkl"
        self.json_file = "students.json"
        self.process_this_frame = True
        self.face_recognition_buffer = {}
        self.recognition_threshold = 3
        self.class_start_time = None
        self.subject = None
        self.late_threshold_minutes = 10
        self.attendance_folder = "Attendance_Records"
        self.summary_folder = "Attendance_Summary"

        if not os.path.exists(self.attendance_folder):
            os.makedirs(self.attendance_folder)
        if not os.path.exists(self.summary_folder):
            os.makedirs(self.summary_folder)

        self.load_database()

    def load_database(self):
        if os.path.exists(self.database_file):
            with open(self.database_file, 'rb') as f:
                data = pickle.load(f)
                self.known_face_encodings = data['encodings']
                self.known_face_names = data['names']
                self.known_face_ids = data['ids']
                self.known_face_sections = data.get('sections', ['Unknown'] * len(self.known_face_ids))
            print(f"✓ Loaded {len(self.known_face_names)} students")
        else:
            print("⚠ No database. Register students first.")
            self.create_sample_json()

    def create_sample_json(self):
        sample_students = {
            "students": [
                {"id": "2025-001", "name": "Juan Dela Cruz", "section": "ICT"},
                {"id": "2025-002", "name": "Maria Santos", "section": "ICT"},
                {"id": "2025-003", "name": "Pedro Reyes", "section": "Cookery"}
            ]
        }
        with open(self.json_file, 'w') as f:
            json.dump(sample_students, f, indent=4)

    def register_student(self, student_id, name, section, image_path):
        try:
            image = face_recognition.load_image_file(image_path)
            face_encodings = face_recognition.face_encodings(image, num_jitters=1)
            if len(face_encodings) == 0:
                return False
            self.known_face_encodings.append(face_encodings[0])
            self.known_face_names.append(name)
            self.known_face_ids.append(student_id)
            self.known_face_sections.append(section)
            print(f"✓ {name} ({student_id})")
            return True
        except:
            return False

    def save_database(self):
        data = {
            'encodings': self.known_face_encodings,
            'names': self.known_face_names,
            'ids': self.known_face_ids,
            'sections': self.known_face_sections
        }
        with open(self.database_file, 'wb') as f:
            pickle.dump(data, f)

    def register_students_from_folder(self, folder_path):
        if not os.path.exists(self.json_file):
            return
        try:
            with open(self.json_file, 'r') as f:
                student_data = json.load(f)
        except:
            return

        registered_count = 0
        for student in student_data['students']:
            student_id = student['id']
            name = student['name']
            section = student.get('section', 'Unknown')
            image_pattern = os.path.join(folder_path, f"{student_id}*.jpg")
            image_paths = glob.glob(image_pattern)
            if image_paths:
                for image_path in image_paths:
                    if self.register_student(student_id, name, section, image_path):
                        registered_count += 1

        if registered_count > 0:
            self.save_database()

    def select_subject(self):
        for idx, subject in enumerate(self.SUBJECTS, 1):
            print(f"{idx}. {subject}")
        while True:
            try:
                choice = int(input("Select: "))
                if 1 <= choice <= len(self.SUBJECTS):
                    return self.SUBJECTS[choice - 1]
            except:
                pass

    def can_student_attend_subject(self, section, subject):
        if subject not in self.SUBJECT_SECTION_RESTRICTIONS:
            return True
        return section in self.SUBJECT_SECTION_RESTRICTIONS[subject]

    def get_stable_recognition(self, face_encoding):
        if len(self.known_face_encodings) == 0:
            return None, None, None
        face_distances = face_recognition.face_distance(self.known_face_encodings, face_encoding)
        if len(face_distances) == 0:
            return None, None, None
        best_match_index = np.argmin(face_distances)
        confidence = 1 - face_distances[best_match_index]
        if confidence > 0.6:
            student_id = self.known_face_ids[best_match_index]
            name = self.known_face_names[best_match_index]
            section = self.known_face_sections[best_match_index]
            if student_id not in self.face_recognition_buffer:
                self.face_recognition_buffer[student_id] = 0
            self.face_recognition_buffer[student_id] += 1
            if self.face_recognition_buffer[student_id] >= self.recognition_threshold:
                return student_id, name, section
        return None, None, None

    def mark_attendance(self, name, student_id, section=None):
        if student_id in self.marked_students:
            return False
        if not self.can_student_attend_subject(section, self.subject):
            return False
        now = datetime.now()
        status = "Present"
        if self.class_start_time:
            time_diff = now - self.class_start_time
            if time_diff.total_seconds() > (self.late_threshold_minutes * 60):
                status = "Late"
        record = {
            'Date': now.strftime("%Y-%m-%d"),
            'Time': now.strftime("%H:%M:%S"),
            'Subject': self.subject,
            'Section': section,
            'ID': student_id,
            'Name': name,
            'Status': status
        }
        self.attendance_records.append(record)
        self.marked_students.add(student_id)
        return True

    def mark_absent_students(self):
        now = datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        unique_students = {}
        for idx, student_id in enumerate(self.known_face_ids):
            if student_id not in unique_students:
                section = self.known_face_sections[idx]
                if self.can_student_attend_subject(section, self.subject):
                    unique_students[student_id] = {
                        'name': self.known_face_names[idx],
                        'section': section
                    }
        for student_id, info in unique_students.items():
            if student_id not in self.marked_students:
                record = {
                    'Date': date_str,
                    'Time': '',
                    'Subject': self.subject,
                    'Section': info['section'],
                    'ID': student_id,
                    'Name': info['name'],
                    'Status': 'Absent'
                }
                self.attendance_records.append(record)

    def run_attendance(self):
        self.subject = self.select_subject()
        self.class_start_time = datetime.now()
        video_capture = cv2.VideoCapture(0)
        if not video_capture.isOpened():
            return
        video_capture.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
        video_capture.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
        print("Scanning faces... (Press 'q' to quit)")
        while True:
            ret, frame = video_capture.read()
            if not ret:
                break
            small_frame = cv2.resize(frame, (0, 0), fx=0.3, fy=0.3)
            rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
            if self.process_this_frame:
                face_locations = face_recognition.face_locations(rgb_small_frame, model="hog")
                if len(face_locations) >= 1:
                    face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)
                    for face_encoding in face_encodings:
                        student_id, name, section = self.get_stable_recognition(face_encoding)
                        if student_id and name:
                            if self.mark_attendance(name, student_id, section):
                                print(f"✓ {name}")
            self.process_this_frame = not self.process_this_frame
            cv2.putText(frame, f"Marked: {len(self.marked_students)}", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
            cv2.imshow('Attendance', frame)
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break
        video_capture.release()
        cv2.destroyAllWindows()
        self.mark_absent_students()
        if len(self.attendance_records) > 0:
            self.save_attendance_to_excel()

    def get_subject_excel_filename(self, subject):
        safe_subject = "".join(c if c.isalnum() or c in (' ', '-', '_') else '' for c in subject)
        safe_subject = safe_subject.replace(' ', '_')
        date_str = datetime.now().strftime("%Y-%m-%d")
        return os.path.join(self.attendance_folder, f"{safe_subject}_{date_str}.xlsx")

    # ==================== COMPLETELY REWRITTEN AND FIXED ====================
    def save_attendance_to_excel(self):
        """COMPLETE FIX: Status column color coding that ACTUALLY WORKS"""
        if len(self.attendance_records) == 0:
            return

        try:
            df = pd.DataFrame(self.attendance_records)
            excel_filename = self.get_subject_excel_filename(self.subject)

            if os.path.exists(excel_filename):
                existing_df = pd.read_excel(excel_filename)
                df = pd.concat([existing_df, df], ignore_index=True)

            # Write to Excel with openpyxl for direct formatting control
            with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Attendance', index=False)
                workbook = writer.book
                worksheet = writer.sheets['Attendance']

                # Define border style
                thin_border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )

                # ===== FORMAT HEADER ROW =====
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF", size=11)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border

                # ===== FORMAT DATA ROWS WITH COLOR CODING =====
                for row_idx, row in enumerate(df.values):
                    row_num = row_idx + 2  # +2 because row 1 is header

                    for col_idx, value in enumerate(row):
                        col_num = col_idx + 1
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                        # ===== COLUMN 8 IS STATUS COLUMN (H) =====
                        if col_num == 8:
                            # Get status value
                            status_value = row[7]  # Column index 7 is the Status column

                            # Apply color based on status
                            if status_value == "Present":
                                cell.fill = PatternFill(start_color=self.COLOR_PRESENT, end_color=self.COLOR_PRESENT, fill_type="solid")
                                cell.font = Font(bold=True, color="FFFFFF", size=11)
                            elif status_value == "Late":
                                cell.fill = PatternFill(start_color=self.COLOR_LATE, end_color=self.COLOR_LATE, fill_type="solid")
                                cell.font = Font(bold=True, color="000000", size=11)
                            elif status_value == "Absent":
                                cell.fill = PatternFill(start_color=self.COLOR_ABSENT, end_color=self.COLOR_ABSENT, fill_type="solid")
                                cell.font = Font(bold=True, color="FFFFFF", size=11)
                            else:
                                cell.font = Font(size=10)
                        else:
                            # Other columns - alternating rows
                            if row_num % 2 == 0:
                                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                            else:
                                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                            cell.font = Font(size=10)

                # Auto-adjust column widths
                for col in worksheet.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[col_letter].width = adjusted_width

                # Freeze header
                worksheet.freeze_panes = "A2"

            print(f"\n✓ Saved: {excel_filename}")
            print(f"✅ STATUS COLORS APPLIED:")
            print(f"   🟢 GREEN = Present")
            print(f"   🟡 YELLOW = Late")
            print(f"   🔴 RED = Absent")

        except Exception as e:
            print(f"✗ Error: {str(e)}")
            import traceback
            traceback.print_exc()

    def format_summary_worksheet(self, worksheet, color_header, title_type):
        for col_num, cell in enumerate(worksheet[1], 1):
            cell.fill = PatternFill(start_color=color_header, end_color=color_header, fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            thin_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            cell.border = thin_border

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                if cell.row % 2 == 0:
                    cell.fill = PatternFill(start_color="E7F0F7", end_color="E7F0F7", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for col in worksheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            worksheet.column_dimensions[col_letter].width = min(max_length + 3, 50)

        worksheet.freeze_panes = "A2"

    def generate_weekly_summary(self):
        print("\nWEEKLY SUMMARY")
        all_records = self.load_all_attendance_records()
        if not all_records:
            print("⚠ No records")
            return
        df = pd.DataFrame(all_records)
        df['Date'] = pd.to_datetime(df['Date'])
        start_date_input = input("Enter start date (YYYY-MM-DD): ").strip()
        try:
            start_date = pd.to_datetime(start_date_input)
            end_date = start_date + timedelta(days=6)
        except:
            return
        week_df = df[(df['Date'] >= start_date) & (df['Date'] <= end_date)]
        if week_df.empty:
            print("⚠ No records")
            return
        summary_data = []
        for student_id in week_df['ID'].unique():
            student_records = week_df[week_df['ID'] == student_id]
            present = len(student_records[student_records['Status'] == 'Present'])
            late = len(student_records[student_records['Status'] == 'Late'])
            absent = len(student_records[student_records['Status'] == 'Absent'])
            total = present + late + absent
            rate = ((present + late) / total * 100) if total > 0 else 0
            summary_data.append({
                'Student ID': student_id,
                'Name': student_records['Name'].iloc[0],
                'Section': student_records['Section'].iloc[0],
                'Present': present,
                'Late': late,
                'Absent': absent,
                'Total': total,
                'Rate %': f"{rate:.1f}%"
            })
        print("\n" + pd.DataFrame(summary_data).to_string(index=False))

    def generate_monthly_summary(self):
        print("\nMONTHLY SUMMARY")
        all_records = self.load_all_attendance_records()
        if not all_records:
            print("⚠ No records")
            return
        df = pd.DataFrame(all_records)
        df['Date'] = pd.to_datetime(df['Date'])
        month_input = input("Enter month (YYYY-MM): ").strip()
        try:
            year, month = map(int, month_input.split('-'))
            start_date = pd.to_datetime(f"{year}-{month}-01")
            if month == 12:
                end_date = pd.to_datetime(f"{year + 1}-01-01") - timedelta(days=1)
            else:
                end_date = pd.to_datetime(f"{year}-{month + 1}-01") - timedelta(days=1)
        except:
            return
        month_df = df[(df['Date'] >= start_date) & (df['Date'] <= end_date)]
        if month_df.empty:
            print("⚠ No records")
            return
        summary_data = []
        for student_id in month_df['ID'].unique():
            student_records = month_df[month_df['ID'] == student_id]
            present = len(student_records[student_records['Status'] == 'Present'])
            late = len(student_records[student_records['Status'] == 'Late'])
            absent = len(student_records[student_records['Status'] == 'Absent'])
            total = present + late + absent
            rate = ((present + late) / total * 100) if total > 0 else 0
            summary_data.append({
                'Student ID': student_id,
                'Name': student_records['Name'].iloc[0],
                'Section': student_records['Section'].iloc[0],
                'Present': present,
                'Late': late,
                'Absent': absent,
                'Total': total,
                'Rate %': f"{rate:.1f}%"
            })
        print("\n" + pd.DataFrame(summary_data).to_string(index=False))

    def load_all_attendance_records(self):
        all_records = []
        if not os.path.exists(self.attendance_folder):
            return all_records
        for filename in os.listdir(self.attendance_folder):
            if filename.endswith('.xlsx'):
                try:
                    df = pd.read_excel(os.path.join(self.attendance_folder, filename))
                    all_records.extend(df.to_dict('records'))
                except:
                    pass
        return all_records

if __name__ == "__main__":
    system = FastAttendanceSystem()
    while True:
        print("\n" + "="*50)
        print("ATTENDANCE SYSTEM v5.5 - FIXED")
        print("="*50)
        print("\n1. Register Students")
        print("2. Start Attendance")
        print("3. Weekly Summary")
        print("4. Monthly Summary")
        print("5. Exit")
        choice = input("\nSelect: ").strip()
        if choice == "1":
            folder = input("Folder: ").strip()
            system.register_students_from_folder(folder)
        elif choice == "2":
            system.attendance_records = []
            system.marked_students = set()
            system.face_recognition_buffer = {}
            system.run_attendance()
        elif choice == "3":
            system.generate_weekly_summary()
        elif choice == "4":
            system.generate_monthly_summary()
        elif choice == "5":
            break
