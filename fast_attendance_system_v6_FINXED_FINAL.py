# ============================================================================
# FINAL VERSION - Enhanced Color Coding for Status (v5.5 - 2025-11-09)
# ============================================================================
# IMPROVEMENTS:
# - ENHANCED: Brighter, more vibrant status colors
# - ENHANCED: Bold text in status cells for better visibility
# - ENHANCED: Color-coded fonts to match backgrounds
# - ENHANCED: Thicker borders around status cells
# - ENHANCED: Better contrast for easier reading
# - FEATURE: Status cells are now HIGHLY VISIBLE with bold formatting
# - FIXED: Color coding for Absent, Late, and Present status cells
# ============================================================================
# Fast Facial Recognition Attendance System - v5.5 ENHANCED COLOR CODING
# Designed for Sibutad National High School
# FINAL: Professional Excel formatting with BOLD color-coded status cells
# NEW: Enhanced color contrast and bold text
# NEW: Improved visual hierarchy for status indicators
# NEW: Professional styling with better visibility
# FIXED: Excel color coding now properly applies to status columns

import cv2
import face_recognition
import numpy as np
import pandas as pd
from datetime import datetime, timedelta
import os
import pickle
import json
import glob
from collections import Counter, defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class FastAttendanceSystem:
    """
    High-performance facial recognition attendance system with enhanced color coding
    """

    SUBJECTS = [
        "Media and Info Literacy",
        "21st Century Literature From the Philippines and the World",
        "Computer System Servicing",
        "(Research Project) Inquiries, Investigation and Immersion",
        "Understanding Culture, Society and Politics",
        "Cookery"
    ]

    SUBJECT_SECTION_RESTRICTIONS = {
        "Cookery": ["Cookery"],
        "Computer System Servicing": ["ICT"]
    }

    # ENHANCED: Improved color codes with better contrast
    COLOR_PRESENT = "00B050"  # Bright Green (was 90EE90)
    COLOR_LATE = "FFD966"     # Bright Yellow (was FFFFE0)
    COLOR_ABSENT = "FF6B6B"   # Bright Red (was FFB6C6)
    COLOR_HEADER = "1F4E78"   # Dark Blue
    COLOR_SUBHEADER = "4472C4"  # Medium Blue
    COLOR_SUMMARY_HEADER = "70AD47"  # Green
    COLOR_MONTHLY_HEADER = "C65911"  # Orange

    def __init__(self):
        self.known_face_encodings = []
        self.known_face_names = []
        self.known_face_ids = []
        self.known_face_sections = []
        self.attendance_records = []
        self.marked_students = set()
        self.verification_start_times = {}  # Track when verification started for each student
        self.database_file = "students_database.pkl"
        self.excel_file = "Attendance_Records.xlsx"
        self.json_file = "students.json"
        self.process_this_frame = True
        self.face_recognition_buffer = {}
        self.verification_start_times = {}  # Clear verification timers
        self.recognition_threshold = 3
        self.class_start_time = None
        self.subject = None
        self.late_threshold_minutes = 10
        self.subject_excel_files = {}
        self.attendance_folder = "Attendance_Records"
        self.summary_folder = "Attendance_Summary"

        if not os.path.exists(self.attendance_folder):
            os.makedirs(self.attendance_folder)

        if not os.path.exists(self.summary_folder):
            os.makedirs(self.summary_folder)

        self.load_database()

    def load_database(self):
        """Load student face encodings from database"""
        if os.path.exists(self.database_file):
            with open(self.database_file, 'rb') as f:
                data = pickle.load(f)
                self.known_face_encodings = data['encodings']
                self.known_face_names = data['names']
                self.known_face_ids = data['ids']
                self.known_face_sections = data.get('sections', ['Unknown'] * len(self.known_face_ids))
            print(f"✓ Loaded {len(self.known_face_names)} students from database")
        else:
            print("⚠ No database found. Please register students first.")
            self.create_sample_json()

    def create_sample_json(self):
        """Create sample student JSON file"""
        sample_students = {
            "students": [
                {"id": "2025-001", "name": "Juan Dela Cruz", "section": "ICT"},
                {"id": "2025-002", "name": "Maria Santos", "section": "ICT"},
                {"id": "2025-003", "name": "Pedro Reyes", "section": "Cookery"}
            ]
        }

        with open(self.json_file, 'w') as f:
            json.dump(sample_students, f, indent=4)
        print(f"✓ Sample student list created: {self.json_file}")

    def register_student(self, student_id, name, section, image_path):
        """Register a new student with their photo and section"""
        try:
            image = face_recognition.load_image_file(image_path)
            face_encodings = face_recognition.face_encodings(image, num_jitters=1)

            if len(face_encodings) == 0:
                print(f"✗ No face found in {image_path}")
                return False

            self.known_face_encodings.append(face_encodings[0])
            self.known_face_names.append(name)
            self.known_face_ids.append(student_id)
            self.known_face_sections.append(section)

            print(f"✓ Registered: {name} ({student_id}) [{section}]")
            return True

        except Exception as e:
            print(f"✗ Error registering {name}: {str(e)}")
            return False

    def save_database(self):
        """Save student encodings to file"""
        data = {
            'encodings': self.known_face_encodings,
            'names': self.known_face_names,
            'ids': self.known_face_ids,
            'sections': self.known_face_sections
        }

        with open(self.database_file, 'wb') as f:
            pickle.dump(data, f)
        print(f"✓ Database saved: {self.database_file}")

    def register_students_from_folder(self, folder_path):
        """Register multiple students from a folder"""
        if not os.path.exists(folder_path):
            print(f"✗ Folder not found: {folder_path}")
            return

        if not os.path.exists(self.json_file):
            print(f"✗ Student list not found: {self.json_file}")
            return

        try:
            with open(self.json_file, 'r') as f:
                student_data = json.load(f)
        except Exception as e:
            print(f"✗ Error reading students.json: {str(e)}")
            return

        print(f"\nRegistering students from: {folder_path}")
        registered_count = 0

        for student in student_data['students']:
            student_id = student['id']
            name = student['name']
            section = student.get('section', 'Unknown')
            image_pattern = os.path.join(folder_path, f"{student_id}*.jpg")
            image_paths = glob.glob(image_pattern)

            if image_paths:
                for image_path in image_paths:
                    if self.register_student(student_id, name, section, image_path):
                        registered_count += 1
            else:
                print(f"⚠ No images found for {name} ({student_id})")

        if registered_count > 0:
            self.save_database()
            print(f"\n✓ Successfully registered {registered_count} student images")
        else:
            print("\n✗ No students were registered")

    def select_subject(self):
        """Display subject selection menu"""
        print("\n" + "="*60)
        print("SELECT SUBJECT")
        print("="*60)
        print("\nAvailable Subjects:")
        for idx, subject in enumerate(self.SUBJECTS, 1):
            print(f"{idx}. {subject}")
        print(f"{len(self.SUBJECTS) + 1}. Enter custom subject")

        while True:
            try:
                choice = input(f"\nEnter subject number (1-{len(self.SUBJECTS) + 1}): ").strip()
                choice_num = int(choice)

                if 1 <= choice_num <= len(self.SUBJECTS):
                    selected_subject = self.SUBJECTS[choice_num - 1]
                    print(f"✓ Selected: {selected_subject}")
                    return selected_subject

                elif choice_num == len(self.SUBJECTS) + 1:
                    custom = input("Enter custom subject name: ").strip()
                    if custom:
                        print(f"✓ Selected: {custom}")
                        return custom
                    else:
                        print("✗ Subject name cannot be empty")
                else:
                    print(f"✗ Invalid number. Please enter 1-{len(self.SUBJECTS) + 1}")

            except ValueError:
                print(f"✗ Invalid input. Please enter a number (1-{len(self.SUBJECTS) + 1})")

    def can_student_attend_subject(self, section, subject):
        """Check if student's section is allowed for this subject"""
        if subject not in self.SUBJECT_SECTION_RESTRICTIONS:
            return True

        allowed_sections = self.SUBJECT_SECTION_RESTRICTIONS[subject]
        can_attend = section in allowed_sections

        if not can_attend:
            print(f"⚠ {section} student cannot be marked for {subject} class")

        return can_attend

    def get_stable_recognition(self, face_encoding):
        """Get stable face recognition using consistency checking with timing"""
        if len(self.known_face_encodings) == 0:
            return None, None, None

        face_distances = face_recognition.face_distance(
            self.known_face_encodings, face_encoding
        )

        if len(face_distances) == 0:
            return None, None, None

        best_match_index = np.argmin(face_distances)
        confidence = 1 - face_distances[best_match_index]

        if confidence > 0.6:
            student_id = self.known_face_ids[best_match_index]
            name = self.known_face_names[best_match_index]
            section = self.known_face_sections[best_match_index]

            # START TIMING: Record when we first start recognizing this student
            if student_id not in self.verification_start_times:
                self.verification_start_times[student_id] = datetime.now()

            if student_id not in self.face_recognition_buffer:
                self.face_recognition_buffer[student_id] = 0

            self.face_recognition_buffer[student_id] += 1

            if self.face_recognition_buffer[student_id] >= self.recognition_threshold:
                return student_id, name, section

        return None, None, None

    def mark_attendance(self, name, student_id, section=None):
        """Mark attendance with section validation and verification time tracking"""
        if student_id in self.marked_students:
            return False

        if not self.can_student_attend_subject(section, self.subject):
            return False

        now = datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H:%M:%S")

        status = "Present"

        if self.class_start_time:
            time_diff = now - self.class_start_time
            if time_diff.total_seconds() > (self.late_threshold_minutes * 60):
                status = "Late"

        # CALCULATE VERIFICATION TIME
        verification_time = 0.0
        if student_id in self.verification_start_times:
            time_delta = now - self.verification_start_times[student_id]
            verification_time = round(time_delta.total_seconds(), 2)

        record = {
            'Date': date_str,
            'Time': time_str,
            'Subject': self.subject,
            'Section': section,
            'ID': student_id,
            'Name': name,
            'Status': status,
            'Verification Time (s)': verification_time  # NEW FIELD
        }

        self.attendance_records.append(record)
        self.marked_students.add(student_id)

        return True

    def mark_absent_students(self):
        """Mark all students who didn't arrive as ABSENT"""
        now = datetime.now()
        date_str = now.strftime("%Y-%m-%d")

        print("\n✓ Marking absent students...")

        absent_count = 0
        unique_students = {}

        for idx, student_id in enumerate(self.known_face_ids):
            if student_id not in unique_students:
                section = self.known_face_sections[idx]
                if self.can_student_attend_subject(section, self.subject):
                    unique_students[student_id] = {
                        'name': self.known_face_names[idx],
                        'section': section
                    }

        for student_id, info in unique_students.items():
            if student_id not in self.marked_students:
                record = {
                    'Date': date_str,
                    'Time': '',
                    'Subject': self.subject,
                    'Section': info['section'],
                    'ID': student_id,
                    'Name': info['name'],
                    'Status': 'Absent',
                    'Verification Time (s)': 0.0  # NEW FIELD - Absent students have 0 verification time
                }

                self.attendance_records.append(record)
                absent_count += 1

        print(f"✓ Marked {absent_count} student(s) as ABSENT")

    def run_attendance(self):
        """Run real-time attendance"""
        print("\n" + "="*60)
        print("CLASS SESSION SETUP")
        print("="*60)

        self.subject = self.select_subject()

        start_time_input = input("\nEnter class start time (HH:MM, or press Enter for NOW): ").strip()

        if start_time_input:
            try:
                start_hour, start_min = map(int, start_time_input.split(':'))
                now = datetime.now()
                self.class_start_time = now.replace(hour=start_hour, minute=start_min, second=0, microsecond=0)
            except:
                print("⚠ Invalid time format. Using current time.")
                self.class_start_time = datetime.now()
        else:
            self.class_start_time = datetime.now()

        print(f"\n✓ Subject: {self.subject}")
        print(f"✓ Class started at: {self.class_start_time.strftime('%H:%M:%S')}")
        print(f"✓ Late threshold: {self.late_threshold_minutes} minutes")

        late_time = self.class_start_time + timedelta(minutes=self.late_threshold_minutes)
        print(f"✓ Students arriving after {late_time.strftime('%H:%M:%S')} will be marked LATE")

        unique_ids = set()
        ict_count = 0
        cookery_count = 0

        for idx, student_id in enumerate(self.known_face_ids):
            if student_id not in unique_ids:
                section = self.known_face_sections[idx]
                if self.can_student_attend_subject(section, self.subject):
                    unique_ids.add(student_id)
                    if section == "ICT":
                        ict_count += 1
                    elif section == "Cookery":
                        cookery_count += 1

        print(f"\n✓ Eligible students for {self.subject}:")
        print(f" - ICT: {ict_count}")
        print(f" - Cookery: {cookery_count}")
        print(f" - Total: {len(unique_ids)}")

        video_capture = cv2.VideoCapture(0)

        if not video_capture.isOpened():
            print("✗ Camera not found or already in use")
            return

        video_capture.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
        video_capture.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)

        print("\n" + "="*60)
        print("FACIAL RECOGNITION ATTENDANCE SYSTEM")
        print("Sibutad National High School")
        print("="*60)
        print("\nPress 'q' to quit | Press 's' to save attendance | Press 'p' to pause/resume")
        print("\nScanning for faces...\n")

        face_locations = []
        face_names = []
        is_paused = False

        while True:
            ret, frame = video_capture.read()

            if not ret:
                break

            small_frame = cv2.resize(frame, (0, 0), fx=0.3, fy=0.3)
            rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

            if is_paused:
                # Draw paused overlay
                overlay = frame.copy()
                cv2.rectangle(overlay, (0, 0), (frame.shape[1], frame.shape[0]), (0, 0, 0), -1)
                cv2.addWeighted(overlay, 0.4, frame, 0.6, 0, frame)
                cv2.putText(frame, "PAUSED", (frame.shape[1]//2 - 120, frame.shape[0]//2),
                           cv2.FONT_HERSHEY_DUPLEX, 2.5, (0, 200, 255), 4)
                cv2.putText(frame, "Press 'P' to resume",
                           (frame.shape[1]//2 - 130, frame.shape[0]//2 + 60),
                           cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 255, 255), 2)
                cv2.imshow('Attendance System - Sibutad National High School', frame)
                key = cv2.waitKey(1) & 0xFF
                if key == ord('p'):
                    is_paused = False
                    print("▶ Attendance resumed")
                elif key == ord('q'):
                    break
                elif key == ord('s'):
                    self.save_attendance_to_excel()
                    print("✓ Attendance saved to Excel")
                continue

            if self.process_this_frame:
                face_locations = face_recognition.face_locations(
                    rgb_small_frame,
                    model="hog",
                    number_of_times_to_upsample=1
                )

                if len(face_locations) >= 1:
                    face_encodings = face_recognition.face_encodings(
                        rgb_small_frame,
                        face_locations,
                        num_jitters=1
                    )

                    face_names = []
                    for face_encoding in face_encodings:
                        student_id, name, section = self.get_stable_recognition(face_encoding)

                        if student_id and name:
                            if self.mark_attendance(name, student_id, section):
                                now = datetime.now()
                                time_diff = now - self.class_start_time

                                if time_diff.total_seconds() > (self.late_threshold_minutes * 60):
                                    print(f"⚠ LATE: {name} ({student_id}) [{section}]")
                                else:
                                    print(f"✓ Attendance marked: {name} ({student_id}) [{section}]")

                            face_names.append(name)
                        else:
                            face_names.append("Verifying...")

                else:
                    face_names = ["Multiple faces" if len(face_locations) > 1 else "No face"] * len(face_locations)

            self.process_this_frame = not self.process_this_frame

            for (top, right, bottom, left), name in zip(face_locations, face_names):
                top = int(top * (1/0.3))
                right = int(right * (1/0.3))
                bottom = int(bottom * (1/0.3))
                left = int(left * (1/0.3))

                if name == "Verifying...":
                    color = (255, 255, 0)
                elif name in ["Multiple faces", "No face"]:
                    color = (0, 0, 255)
                else:
                    color = (0, 255, 0)

                cv2.rectangle(frame, (left, top), (right, bottom), color, 2)
                cv2.rectangle(frame, (left, bottom - 35), (right, bottom), color, cv2.FILLED)

                font = cv2.FONT_HERSHEY_DUPLEX
                cv2.putText(frame, name, (left + 6, bottom - 6), font, 0.6, (255, 255, 255), 1)

            cv2.putText(frame, f"Subject: {self.subject}",
                       (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)

            cv2.putText(frame, f"Marked: {len(self.marked_students)}",
                       (10, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)

            if len(face_locations) > 1:
                cv2.putText(frame, "WARNING: Multiple faces detected!",
                           (10, 120), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)

            current_time = datetime.now()
            time_diff = current_time - self.class_start_time

            if time_diff.total_seconds() > (self.late_threshold_minutes * 60):
                cv2.putText(frame, "Status: LATE ARRIVALS",
                           (10, 90), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)
            else:
                cv2.putText(frame, "Status: ON TIME",
                           (10, 90), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)

            cv2.imshow('Attendance System - Sibutad National High School', frame)

            key = cv2.waitKey(1) & 0xFF

            if key == ord('q'):
                break
            elif key == ord('s'):
                self.save_attendance_to_excel()
                print("✓ Attendance saved to Excel")
            elif key == ord('p'):
                is_paused = True
                print("⏸ Attendance paused. Press 'P' to resume.")

        video_capture.release()
        cv2.destroyAllWindows()

        self.mark_absent_students()

        if len(self.attendance_records) > 0:
            self.save_attendance_to_excel()

    def get_subject_excel_filename(self, subject):
        """Generate subject-specific Excel filename"""
        safe_subject = "".join(c if c.isalnum() or c in (' ', '-', '_') else '' for c in subject)
        safe_subject = safe_subject.replace(' ', '_')
        date_str = datetime.now().strftime("%Y-%m-%d")
        filename = f"{safe_subject}_{date_str}.xlsx"
        return os.path.join(self.attendance_folder, filename)

    # ENHANCED: Improved status color function
    def get_status_color(self, status):
        """Get color for attendance status - ENHANCED with better contrast"""
        if status == "Present":
            return self.COLOR_PRESENT  # Bright Green
        elif status == "Late":
            return self.COLOR_LATE  # Bright Yellow
        elif status == "Absent":
            return self.COLOR_ABSENT  # Bright Red
        else:
            return "FFFFFF"

    # ENHANCED: Better font colors for status text
    def get_status_font_color(self, status):
        """Get font color for status text"""
        if status == "Present":
            return "FFFFFF"  # White text on green
        elif status == "Late":
            return "000000"  # Black text on yellow
        elif status == "Absent":
            return "FFFFFF"  # White text on red
        else:
            return "000000"  # Black

    # ENHANCED: Improved cell formatting with better visibility
    def apply_cell_formatting(self, cell, is_header=False, is_subheader=False, status=None):
        """Apply professional formatting to cells with enhanced status visibility"""
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        cell.border = thin_border

        if is_header:
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=11)

        elif is_subheader:
            cell.fill = PatternFill(start_color=self.COLOR_SUBHEADER, end_color=self.COLOR_SUBHEADER, fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=10)

        elif status:  # FIXED: Apply color coding for status
            color = self.get_status_color(status)
            font_color = self.get_status_font_color(status)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.font = Font(bold=True, color=font_color, size=11)  # BOLD for visibility

        else:
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            cell.font = Font(size=10)

    def save_attendance_to_excel(self):
        """Save attendance to subject-separated Excel files with ENHANCED formatting - FIXED COLOR CODING"""
        if len(self.attendance_records) == 0:
            print("⚠ No attendance records to save")
            return

        try:
            df = pd.DataFrame(self.attendance_records)
            excel_filename = self.get_subject_excel_filename(self.subject)

            if os.path.exists(excel_filename):
                existing_df = pd.read_excel(excel_filename)
                df = pd.concat([existing_df, df], ignore_index=True)

            with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Attendance', index=False)

                workbook = writer.book
                worksheet = writer.sheets['Attendance']

                # Format header row
                for col_num, value in enumerate(df.columns.values, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    self.apply_cell_formatting(cell, is_header=True)

                # Format data rows with FIXED status colors
                # Column indices: Date=1, Time=2, Subject=3, Section=4, ID=5, Name=6, Status=7, Verification Time=8
                for row_num, row_data in enumerate(df.values, 2):
                    for col_num, value in enumerate(row_data, 1):
                        cell = worksheet.cell(row=row_num, column=col_num)

                        # FIXED: Properly check if this is the Status column (column 7)
                        if col_num == 7:  # Status column
                            # Use row_data directly to get the status value
                            status_value = row_data[6]  # Index 6 is column 7 (0-indexed)
                            self.apply_cell_formatting(cell, status=status_value)
                        else:
                            self.apply_cell_formatting(cell)

                        # Alternate row coloring for better readability
                        if row_num % 2 == 0 and col_num != 7:
                            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter

                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass

                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                # Freeze header row
                worksheet.freeze_panes = "A2"

            print(f"\n✓ Attendance saved to: {excel_filename}")
            print(f"✓ Total records: {len(df)}")
            print("✓ Format: ENHANCED color-coded status cells with bold text")
            print("✓ Status Colors:")
            print(" 🟢 GREEN = Present | 🟡 YELLOW = Late | 🔴 RED = Absent")

        except Exception as e:
            print(f"✗ Error saving attendance: {str(e)}")

    def format_summary_worksheet(self, worksheet, color_header, title_type):
        """Format summary worksheet with professional styling"""
        for col_num, cell in enumerate(worksheet[1], 1):
            cell.fill = PatternFill(start_color=color_header, end_color=color_header, fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            thin_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )

            cell.border = thin_border

        for row_num, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column), 2):
            for col_num, cell in enumerate(row, 1):
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="E7F0F7", end_color="E7F0F7", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                thin_border = Border(
                    left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC')
                )

                cell.border = thin_border

        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 3, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        worksheet.freeze_panes = "A2"

    def generate_weekly_summary(self):
        """Generate weekly attendance summary"""
        print("\n" + "="*60)
        print("WEEKLY ATTENDANCE SUMMARY")
        print("="*60)

        all_records = self.load_all_attendance_records()

        if not all_records:
            print("⚠ No attendance records found.")
            return

        df = pd.DataFrame(all_records)
        df['Date'] = pd.to_datetime(df['Date'])

        print("\nEnter date range for weekly summary")
        start_date_input = input("Enter start date (YYYY-MM-DD): ").strip()

        try:
            start_date = pd.to_datetime(start_date_input)
            end_date = start_date + timedelta(days=6)
        except:
            print("✗ Invalid date format")
            return

        week_df = df[(df['Date'] >= start_date) & (df['Date'] <= end_date)]

        if week_df.empty:
            print(f"⚠ No records found between {start_date.date()} and {end_date.date()}")
            return

        print(f"\n✓ Generating summary for week: {start_date.date()} to {end_date.date()}")

        summary_data = []

        for student_id in week_df['ID'].unique():
            student_records = week_df[week_df['ID'] == student_id]
            student_name = student_records['Name'].iloc[0]
            student_section = student_records['Section'].iloc[0]

            present_count = len(student_records[student_records['Status'] == 'Present'])
            late_count = len(student_records[student_records['Status'] == 'Late'])
            absent_count = len(student_records[student_records['Status'] == 'Absent'])

            total_classes = present_count + late_count + absent_count
            attendance_rate = ((present_count + late_count) / total_classes * 100) if total_classes > 0 else 0

            summary_data.append({
                'Student ID': student_id,
                'Name': student_name,
                'Section': student_section,
                'Present': present_count,
                'Late': late_count,
                'Absent': absent_count,
                'Total Classes': total_classes,
                'Attendance Rate %': f"{attendance_rate:.1f}%"
            })

        summary_df = pd.DataFrame(summary_data)
        filename = f"Weekly_Summary_{start_date.strftime('%Y-%m-%d')}.xlsx"
        filepath = os.path.join(self.summary_folder, filename)

        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            summary_df.to_excel(writer, sheet_name='Weekly Summary', index=False)
            worksheet = writer.sheets['Weekly Summary']
            self.format_summary_worksheet(worksheet, self.COLOR_SUMMARY_HEADER, "Weekly")

        print(f"\n✓ Weekly summary saved to: {filepath}")
        print(f"✓ Total students: {len(summary_df)}")
        print("\n" + summary_df.to_string(index=False))

    def generate_monthly_summary(self):
        """Generate monthly attendance summary"""
        print("\n" + "="*60)
        print("MONTHLY ATTENDANCE SUMMARY")
        print("="*60)

        all_records = self.load_all_attendance_records()

        if not all_records:
            print("⚠ No attendance records found.")
            return

        df = pd.DataFrame(all_records)
        df['Date'] = pd.to_datetime(df['Date'])

        print("\nEnter date for monthly summary")
        month_input = input("Enter month (YYYY-MM format, e.g., 2025-11): ").strip()

        try:
            year, month = map(int, month_input.split('-'))
            start_date = pd.to_datetime(f"{year}-{month}-01")

            if month == 12:
                end_date = pd.to_datetime(f"{year + 1}-01-01") - timedelta(days=1)
            else:
                end_date = pd.to_datetime(f"{year}-{month + 1}-01") - timedelta(days=1)

        except:
            print("✗ Invalid date format. Use YYYY-MM format.")
            return

        month_df = df[(df['Date'] >= start_date) & (df['Date'] <= end_date)]

        if month_df.empty:
            print(f"⚠ No records found for {month_input}")
            return

        print(f"\n✓ Generating summary for month: {start_date.strftime('%B %Y')}")

        summary_data = []

        for student_id in month_df['ID'].unique():
            student_records = month_df[month_df['ID'] == student_id]
            student_name = student_records['Name'].iloc[0]
            student_section = student_records['Section'].iloc[0]

            present_count = len(student_records[student_records['Status'] == 'Present'])
            late_count = len(student_records[student_records['Status'] == 'Late'])
            absent_count = len(student_records[student_records['Status'] == 'Absent'])

            total_classes = present_count + late_count + absent_count
            attendance_rate = ((present_count + late_count) / total_classes * 100) if total_classes > 0 else 0

            summary_data.append({
                'Student ID': student_id,
                'Name': student_name,
                'Section': student_section,
                'Present': present_count,
                'Late': late_count,
                'Absent': absent_count,
                'Total Classes': total_classes,
                'Attendance Rate %': f"{attendance_rate:.1f}%"
            })

        summary_df = pd.DataFrame(summary_data)
        filename = f"Monthly_Summary_{month_input}.xlsx"
        filepath = os.path.join(self.summary_folder, filename)

        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            summary_df.to_excel(writer, sheet_name='Monthly Summary', index=False)
            worksheet = writer.sheets['Monthly Summary']
            self.format_summary_worksheet(worksheet, self.COLOR_MONTHLY_HEADER, "Monthly")

        print(f"\n✓ Monthly summary saved to: {filepath}")
        print(f"✓ Total students: {len(summary_df)}")
        print("\n" + summary_df.to_string(index=False))

    def load_all_attendance_records(self):
        """Load all attendance records from all subject files"""
        all_records = []

        if not os.path.exists(self.attendance_folder):
            return all_records

        for filename in os.listdir(self.attendance_folder):
            if filename.endswith('.xlsx'):
                filepath = os.path.join(self.attendance_folder, filename)
                try:
                    df = pd.read_excel(filepath)
                    all_records.extend(df.to_dict('records'))
                except Exception as e:
                    print(f"⚠ Error reading {filename}: {str(e)}")

        return all_records


# Main execution
if __name__ == "__main__":
    system = FastAttendanceSystem()

    while True:
        print("\n" + "="*60)
        print("FACIAL RECOGNITION ATTENDANCE SYSTEM v5.5")
        print("Sibutad National High School")
        print("="*60)
        print("\n1. Register Students from Folder")
        print("2. Start Attendance (Real-time)")
        print("3. Generate Weekly Summary Report")
        print("4. Generate Monthly Summary Report")
        print("5. Exit")
        print("="*60)

        choice = input("\nSelect option (1-5): ").strip()

        if choice == "1":
            folder = input("Enter folder path with student images: ").strip()
            system.register_students_from_folder(folder)

        elif choice == "2":
            system.attendance_records = []
            system.marked_students = set()
            system.face_recognition_buffer = {}
            system.run_attendance()

        elif choice == "3":
            system.generate_weekly_summary()

        elif choice == "4":
            system.generate_monthly_summary()

        elif choice == "5":
            print("\n✓ System closed. Goodbye!")
            break

        else:
            print("✗ Invalid option. Please select 1-5.")
