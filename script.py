
# Create a fast facial recognition attendance system with Excel output
# Based on the research paper requirements from Sibutad National High School

code_content = '''# Fast Facial Recognition Attendance System
# Designed for Sibutad National High School
# Based on research: Manual vs. Automated Attendance Monitoring Systems
# Optimized for speed and proper Excel layout

import cv2
import face_recognition
import numpy as np
import pandas as pd
from datetime import datetime
import os
import pickle
import threading
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

class FastAttendanceSystem:
    """
    High-performance facial recognition attendance system
    Features:
    - Fast face detection and recognition
    - Properly formatted Excel output
    - Real-time processing
    - Organized data storage
    """
    
    def __init__(self):
        self.known_face_encodings = []
        self.known_face_names = []
        self.known_face_ids = []
        self.attendance_records = []
        self.database_file = "students_database.pkl"
        self.excel_file = "Attendance_Records.xlsx"
        self.json_file = "students.json"
        self.process_this_frame = True
        
        # Load student database
        self.load_database()
    
    def load_database(self):
        """Load student face encodings from database"""
        if os.path.exists(self.database_file):
            with open(self.database_file, 'rb') as f:
                data = pickle.load(f)
                self.known_face_encodings = data['encodings']
                self.known_face_names = data['names']
                self.known_face_ids = data['ids']
            print(f"✓ Loaded {len(self.known_face_names)} students from database")
        else:
            print("⚠ No database found. Please register students first.")
            self.create_sample_json()
    
    def create_sample_json(self):
        """Create sample student JSON file"""
        sample_students = {
            "students": [
                {"id": "2024-001", "name": "Juan Dela Cruz", "section": "Grade 12 ICT"},
                {"id": "2024-002", "name": "Maria Santos", "section": "Grade 12 ICT"},
                {"id": "2024-003", "name": "Pedro Reyes", "section": "Grade 12 ICT"}
            ]
        }
        with open(self.json_file, 'w') as f:
            json.dump(sample_students, f, indent=4)
        print(f"✓ Sample student list created: {self.json_file}")
    
    def register_student(self, student_id, name, section, image_path):
        """Register a new student with their photo"""
        # Load image
        image = face_recognition.load_image_file(image_path)
        
        # Get face encoding (fast method)
        face_encodings = face_recognition.face_encodings(image, num_jitters=1)
        
        if len(face_encodings) == 0:
            print(f"✗ No face found in {image_path}")
            return False
        
        # Add to database
        self.known_face_encodings.append(face_encodings[0])
        self.known_face_names.append(name)
        self.known_face_ids.append(student_id)
        
        print(f"✓ Registered: {name} ({student_id})")
        return True
    
    def save_database(self):
        """Save student encodings to file"""
        data = {
            'encodings': self.known_face_encodings,
            'names': self.known_face_names,
            'ids': self.known_face_ids
        }
        with open(self.database_file, 'wb') as f:
            pickle.dump(data, f)
        print(f"✓ Database saved: {self.database_file}")
    
    def register_students_from_folder(self, folder_path):
        """Register multiple students from a folder with JSON metadata"""
        if not os.path.exists(self.json_file):
            print(f"✗ Student list not found: {self.json_file}")
            return
        
        # Load student information
        with open(self.json_file, 'r') as f:
            student_data = json.load(f)
        
        # Register each student
        for student in student_data['students']:
            student_id = student['id']
            name = student['name']
            section = student['section']
            
            # Look for student photo
            image_file = f"{student_id}.jpg"
            image_path = os.path.join(folder_path, image_file)
            
            if os.path.exists(image_path):
                self.register_student(student_id, name, section, image_path)
            else:
                print(f"⚠ Image not found for {name}: {image_path}")
        
        self.save_database()
    
    def mark_attendance(self, name, student_id):
        """Mark attendance for a student"""
        now = datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H:%M:%S")
        
        # Check if already marked today
        for record in self.attendance_records:
            if record['ID'] == student_id and record['Date'] == date_str:
                return False  # Already marked
        
        # Add attendance record
        record = {
            'Date': date_str,
            'Time': time_str,
            'ID': student_id,
            'Name': name,
            'Status': 'Present'
        }
        self.attendance_records.append(record)
        return True
    
    def run_attendance(self):
        """Run real-time attendance with optimized face recognition"""
        video_capture = cv2.VideoCapture(0)
        
        # Set camera properties for speed
        video_capture.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
        video_capture.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
        
        print("\\n" + "="*60)
        print("FACIAL RECOGNITION ATTENDANCE SYSTEM")
        print("Sibutad National High School")
        print("="*60)
        print("\\nPress 'q' to quit | Press 's' to save attendance")
        print("\\nScanning for faces...")
        
        face_locations = []
        face_encodings = []
        face_names = []
        
        while True:
            ret, frame = video_capture.read()
            
            if not ret:
                break
            
            # Resize frame for faster processing (key optimization)
            small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
            rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
            
            # Process every other frame for speed
            if self.process_this_frame:
                # Detect faces (optimized)
                face_locations = face_recognition.face_locations(rgb_small_frame, model="hog")
                face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations, num_jitters=1)
                
                face_names = []
                for face_encoding in face_encodings:
                    # Compare faces (optimized with tolerance)
                    matches = face_recognition.compare_faces(
                        self.known_face_encodings, 
                        face_encoding, 
                        tolerance=0.6
                    )
                    name = "Unknown"
                    student_id = None
                    
                    # Use face distance for better accuracy
                    face_distances = face_recognition.face_distance(
                        self.known_face_encodings, 
                        face_encoding
                    )
                    
                    if len(face_distances) > 0:
                        best_match_index = np.argmin(face_distances)
                        if matches[best_match_index]:
                            name = self.known_face_names[best_match_index]
                            student_id = self.known_face_ids[best_match_index]
                            
                            # Mark attendance
                            if self.mark_attendance(name, student_id):
                                print(f"✓ Attendance marked: {name} ({student_id})")
                    
                    face_names.append(name)
            
            self.process_this_frame = not self.process_this_frame
            
            # Display results
            for (top, right, bottom, left), name in zip(face_locations, face_names):
                # Scale back up face locations
                top *= 4
                right *= 4
                bottom *= 4
                left *= 4
                
                # Draw box
                color = (0, 255, 0) if name != "Unknown" else (0, 0, 255)
                cv2.rectangle(frame, (left, top), (right, bottom), color, 2)
                
                # Draw label
                cv2.rectangle(frame, (left, bottom - 35), (right, bottom), color, cv2.FILLED)
                font = cv2.FONT_HERSHEY_DUPLEX
                cv2.putText(frame, name, (left + 6, bottom - 6), font, 0.6, (255, 255, 255), 1)
            
            # Display info
            cv2.putText(frame, f"Students Registered: {len(self.known_face_names)}", 
                       (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
            cv2.putText(frame, f"Attendance Today: {len(self.attendance_records)}", 
                       (10, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
            
            cv2.imshow('Attendance System - Sibutad National High School', frame)
            
            # Controls
            key = cv2.waitKey(1) & 0xFF
            if key == ord('q'):
                break
            elif key == ord('s'):
                self.save_attendance_to_excel()
                print("✓ Attendance saved to Excel")
        
        video_capture.release()
        cv2.destroyAllWindows()
        
        # Auto-save on exit
        if len(self.attendance_records) > 0:
            self.save_attendance_to_excel()
    
    def save_attendance_to_excel(self):
        """Save attendance to properly formatted Excel file"""
        if len(self.attendance_records) == 0:
            print("⚠ No attendance records to save")
            return
        
        # Create DataFrame
        df = pd.DataFrame(self.attendance_records)
        
        # Sort by date and time
        df = df.sort_values(['Date', 'Time'])
        
        # Create Excel file with formatting
        try:
            if os.path.exists(self.excel_file):
                # Append to existing file
                existing_df = pd.read_excel(self.excel_file)
                df = pd.concat([existing_df, df], ignore_index=True)
                df = df.drop_duplicates(subset=['Date', 'ID'], keep='first')
        except:
            pass
        
        # Save to Excel
        df.to_excel(self.excel_file, index=False, sheet_name='Attendance')
        
        # Format Excel file
        self.format_excel_file()
        
        print(f"\\n✓ Attendance saved: {self.excel_file}")
        print(f"  Total records: {len(df)}")
    
    def format_excel_file(self):
        """Apply professional formatting to Excel file"""
        wb = load_workbook(self.excel_file)
        ws = wb.active
        
        # Set title
        ws.title = "Attendance Records"
        
        # Header formatting
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Apply header formatting
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply borders and alignment to all cells
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Adjust column widths
        column_widths = {
            'A': 15,  # Date
            'B': 12,  # Time
            'C': 15,  # ID
            'D': 25,  # Name
            'E': 12   # Status
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Add title row
        ws.insert_rows(1)
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = "SIBUTAD NATIONAL HIGH SCHOOL - ATTENDANCE RECORD"
        title_cell.font = Font(bold=True, size=14, color="4472C4")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Save workbook
        wb.save(self.excel_file)
    
    def generate_attendance_report(self, start_date=None, end_date=None):
        """Generate attendance summary report"""
        if not os.path.exists(self.excel_file):
            print("⚠ No attendance records found")
            return
        
        df = pd.read_excel(self.excel_file)
        
        # Filter by date range if provided
        if start_date:
            df = df[df['Date'] >= start_date]
        if end_date:
            df = df[df['Date'] <= end_date]
        
        # Generate summary
        print("\\n" + "="*60)
        print("ATTENDANCE SUMMARY REPORT")
        print("="*60)
        print(f"Total Records: {len(df)}")
        print(f"Unique Students: {df['ID'].nunique()}")
        print(f"Date Range: {df['Date'].min()} to {df['Date'].max()}")
        print("\\nAttendance by Student:")
        print(df.groupby(['ID', 'Name']).size().reset_index(name='Days Present'))


# Main execution
if __name__ == "__main__":
    system = FastAttendanceSystem()
    
    print("\\n" + "="*60)
    print("FACIAL RECOGNITION ATTENDANCE SYSTEM")
    print("Sibutad National High School")
    print("Based on Research: Manual vs. Automated Attendance Monitoring")
    print("="*60)
    print("\\nOptions:")
    print("1. Register students from folder")
    print("2. Run attendance")
    print("3. Generate report")
    print("4. Exit")
    
    choice = input("\\nEnter choice (1-4): ")
    
    if choice == "1":
        folder = input("Enter folder path with student photos: ")
        system.register_students_from_folder(folder)
    elif choice == "2":
        system.run_attendance()
    elif choice == "3":
        system.generate_attendance_report()
    elif choice == "4":
        print("Goodbye!")
    else:
        print("Invalid choice")
'''

# Save the code to a file
with open('fast_attendance_system.py', 'w', encoding='utf-8') as f:
    f.write(code_content)

print("✓ Fast Facial Recognition Attendance System created!")
print("✓ File: fast_attendance_system.py")
print("\nFeatures:")
print("- Optimized for speed (processes every other frame)")
print("- Properly formatted Excel output")
print("- Real-time face detection and recognition")
print("- Student database management")
print("- Professional attendance reports")
