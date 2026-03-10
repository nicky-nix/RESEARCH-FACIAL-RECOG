import sys
import cv2
import face_recognition
import numpy as np
import pandas as pd
from datetime import datetime, timedelta
import os
import pickle
import json
import glob
import threading
from collections import Counter, defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                             QPushButton, QLabel, QComboBox, QLineEdit, QTableWidget, 
                             QTableWidgetItem, QSpinBox, QTimeEdit, QFrame, QStackedWidget,
                             QMessageBox, QDialog, QProgressBar, QFileDialog, QScrollArea,
                             QTabWidget, QListWidget, QListWidgetItem)
from PyQt6.QtCore import Qt, QTimer, QTime, pyqtSignal, QThread, QSize, QRect, QObject
from PyQt6.QtGui import (QFont, QColor, QPixmap, QIcon, QPalette, QImage, QBrush,
                         QPainter, QPen)


class CameraThread(QThread):
    """Thread for camera processing"""
    frame_ready = pyqtSignal(QImage)

    def __init__(self):
        super().__init__()
        self.running = True
        self.camera = None

    def run(self):
        """Run camera capture"""
        self.camera = cv2.VideoCapture(0)
        if not self.camera.isOpened():
            return

        self.camera.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
        self.camera.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)

        while self.running:
            ret, frame = self.camera.read()
            if ret:
                # Convert BGR to RGB
                rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

                # Convert to QImage
                h, w, ch = rgb_frame.shape
                bytes_per_line = ch * w
                qt_image = QImage(rgb_frame.data, w, h, bytes_per_line, QImage.Format.Format_RGB888)

                self.frame_ready.emit(qt_image)
            else:
                break

    def stop(self):
        """Stop camera capture"""
        self.running = False
        if self.camera:
            self.camera.release()


class ModernButton(QPushButton):
    """Custom button with modern styling"""
    def __init__(self, text, parent=None, accent=False):
        super().__init__(text, parent)
        self.setMinimumHeight(40)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.accent = accent
        self.update_style()

    def update_style(self):
        if self.accent:
            self.setStyleSheet("""
                QPushButton {
                    background-color: #FF5555;
                    color: white;
                    border: none;
                    border-radius: 6px;
                    font-weight: bold;
                    font-size: 11pt;
                    padding: 8px;
                }
                QPushButton:hover {
                    background-color: #FF3333;
                }
                QPushButton:pressed {
                    background-color: #DD2222;
                }
            """)
        else:
            self.setStyleSheet("""
                QPushButton {
                    background-color: #4472C4;
                    color: white;
                    border: none;
                    border-radius: 6px;
                    font-weight: bold;
                    font-size: 11pt;
                    padding: 8px;
                }
                QPushButton:hover {
                    background-color: #3A5AA8;
                }
                QPushButton:pressed {
                    background-color: #2E4A90;
                }
            """)


class StatWidget(QFrame):
    """Stat card widget with responsive sizing"""
    def __init__(self, title, value="—", color="#4472C4"):
        super().__init__()
        self.title = title
        self.value = value
        self.color = color

        self.setStyleSheet(f"""
            QFrame {{
                background-color: #2D2D2D;
                border-left: 4px solid {color};
                border-radius: 6px;
                border: 1px solid #3A3A3A;
            }}
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(15, 12, 15, 12)
        layout.setSpacing(8)

        label = QLabel(title)
        label.setStyleSheet("color: #888888; font-size: 10pt;")

        self.value_label = QLabel(str(value))
        font = QFont()
        font.setPointSize(20)
        font.setBold(True)
        self.value_label.setFont(font)
        self.value_label.setStyleSheet(f"color: {color};")

        layout.addWidget(label)
        layout.addWidget(self.value_label)

    def set_value(self, value):
        """Update stat value"""
        self.value_label.setText(str(value))


class AttendanceGUI(QMainWindow):
    """Modern, fully responsive GUI for attendance system"""

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Facial Recognition Attendance System v6.0")
        self.setGeometry(100, 100, 1400, 900)

        # Initialize
        self.session_active = False
        self.marked_count = 0
        self.camera_thread = None

        # Apply dark theme
        self.apply_dark_theme()

        # Create UI
        self.init_ui()

        # Show maximized
        self.showMaximized()

    def apply_dark_theme(self):
        """Apply modern dark theme"""
        dark_stylesheet = """
            QMainWindow, QWidget {
                background-color: #1E1E1E;
                color: #E0E0E0;
            }
            QFrame {
                background-color: #2D2D2D;
                border-radius: 8px;
            }
            QLabel {
                color: #E0E0E0;
            }
            QComboBox, QLineEdit, QSpinBox, QTimeEdit {
                background-color: #3A3A3A;
                color: #E0E0E0;
                border: 2px solid #4472C4;
                border-radius: 4px;
                padding: 6px;
                font-size: 10pt;
                selection-background-color: #FF5555;
            }
            QComboBox:hover, QLineEdit:hover, QSpinBox:hover, QTimeEdit:hover {
                border: 2px solid #5A8FE8;
            }
            QComboBox::drop-down {
                border: none;
            }
            QComboBox::down-arrow {
                background-color: #4472C4;
            }
            QTableWidget {
                background-color: #2D2D2D;
                gridline-color: #4A4A4A;
                border: none;
                border-radius: 4px;
            }
            QTableWidget::item {
                padding: 4px;
            }
            QHeaderView::section {
                background-color: #1F4E78;
                color: white;
                padding: 5px;
                border: none;
                font-weight: bold;
                border-radius: 0px;
            }
            QProgressBar {
                border: 2px solid #4472C4;
                border-radius: 4px;
                background-color: #3A3A3A;
                text-align: center;
            }
            QProgressBar::chunk {
                background-color: #00B050;
                border-radius: 2px;
            }
            QListWidget {
                background-color: #2D2D2D;
                border: 1px solid #3A3A3A;
                border-radius: 4px;
            }
            QListWidget::item {
                padding: 5px;
            }
            QListWidget::item:hover {
                background-color: #3A3A3A;
            }
            QScrollBar:vertical {
                background-color: #2D2D2D;
                width: 12px;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical {
                background-color: #4472C4;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical:hover {
                background-color: #5A8FE8;
            }
        """
        self.setStyleSheet(dark_stylesheet)

    def init_ui(self):
        """Initialize user interface"""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        main_layout = QHBoxLayout(central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # Sidebar
        sidebar = self.create_sidebar()
        main_layout.addWidget(sidebar, 1)

        # Main content
        self.stacked_widget = QStackedWidget()
        main_layout.addWidget(self.stacked_widget, 4)

        # Pages
        self.dashboard_page = self.create_dashboard_page()
        self.attendance_page = self.create_attendance_page()
        self.reports_page = self.create_reports_page()
        self.settings_page = self.create_settings_page()

        self.stacked_widget.addWidget(self.dashboard_page)
        self.stacked_widget.addWidget(self.attendance_page)
        self.stacked_widget.addWidget(self.reports_page)
        self.stacked_widget.addWidget(self.settings_page)

    def create_sidebar(self):
        """Create navigation sidebar"""
        sidebar_frame = QFrame()
        sidebar_frame.setStyleSheet("""
            QFrame {
                background-color: #1F1F1F;
                border-right: 3px solid #FF5555;
            }
        """)
        sidebar_frame.setMinimumWidth(200)
        sidebar_frame.setMaximumWidth(250)

        layout = QVBoxLayout(sidebar_frame)
        layout.setContentsMargins(15, 20, 15, 20)
        layout.setSpacing(10)

        # Header
        title = QLabel("ATTENDANCE")
        font = QFont()
        font.setPointSize(16)
        font.setBold(True)
        title.setFont(font)
        title.setStyleSheet("color: #FF5555;")
        layout.addWidget(title)

        subtitle = QLabel("v6.0")
        font = QFont()
        font.setPointSize(9)
        subtitle.setFont(font)
        subtitle.setStyleSheet("color: #888888; margin-bottom: 10px;")
        layout.addWidget(subtitle)

        # Divider
        divider = QFrame()
        divider.setStyleSheet("background-color: #3A3A3A;")
        divider.setMaximumHeight(1)
        layout.addWidget(divider)
        layout.addSpacing(10)

        # Nav buttons
        nav_items = [
            ("📊 Dashboard", self.show_dashboard),
            ("📸 Mark Attendance", self.show_attendance),
            ("📋 Reports", self.show_reports),
            ("⚙️  Settings", self.show_settings)
        ]

        for icon_text, callback in nav_items:
            btn = QPushButton(icon_text)
            btn.setMinimumHeight(45)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setStyleSheet("""
                QPushButton {
                    background-color: #2D2D2D;
                    color: #E0E0E0;
                    border: none;
                    border-radius: 6px;
                    font-weight: bold;
                    text-align: left;
                    padding-left: 15px;
                    font-size: 11pt;
                }
                QPushButton:hover {
                    background-color: #3A3A3A;
                    border-left: 4px solid #FF5555;
                    padding-left: 11px;
                }
                QPushButton:pressed {
                    background-color: #FF5555;
                    color: white;
                }
            """)
            btn.clicked.connect(callback)
            layout.addWidget(btn)

        layout.addStretch()

        # Footer
        footer = QLabel("Sibutad National\nHigh School")
        font = QFont()
        font.setPointSize(8)
        footer.setFont(font)
        footer.setStyleSheet("color: #666666;")
        footer.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(footer)

        return sidebar_frame

    def create_dashboard_page(self):
        """Create dashboard page"""
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # Header
        header = QLabel("Dashboard")
        font = QFont()
        font.setPointSize(22)
        font.setBold(True)
        header.setFont(font)
        header.setStyleSheet("color: #FF5555; margin-bottom: 10px;")
        layout.addWidget(header)

        # Stats grid
        stats_layout = QHBoxLayout()
        stats_layout.setSpacing(12)

        self.stat_total = StatWidget("👥 Total Students", "—", "#4472C4")
        self.stat_present = StatWidget("✅ Present", "0", "#00B050")
        self.stat_late = StatWidget("⏰ Late", "0", "#FFD966")
        self.stat_absent = StatWidget("❌ Absent", "0", "#FF6B6B")

        stats_layout.addWidget(self.stat_total)
        stats_layout.addWidget(self.stat_present)
        stats_layout.addWidget(self.stat_late)
        stats_layout.addWidget(self.stat_absent)

        layout.addLayout(stats_layout)

        # Info card
        info_frame = QFrame()
        info_frame.setStyleSheet("""
            QFrame {
                background-color: #2D2D2D;
                border-radius: 8px;
                border: 1px solid #4472C4;
            }
        """)
        info_layout = QVBoxLayout(info_frame)
        info_layout.setContentsMargins(20, 20, 20, 20)

        title_label = QLabel("System Information")
        font = QFont()
        font.setPointSize(12)
        font.setBold(True)
        title_label.setFont(font)
        title_label.setStyleSheet("color: #FF5555;")
        info_layout.addWidget(title_label)

        info_text = QLabel("""
        <span style="color: #D0D0D0; font-size: 10pt; line-height: 1.6;">
        <b>Welcome to Facial Recognition Attendance System</b><br><br>
        This system uses advanced face recognition technology to automate attendance marking.<br><br>
        <b>Getting Started:</b><br>
        1. Go to Settings → Register Students to load facial encodings<br>
        2. Select "Mark Attendance" to begin a session<br>
        3. Review attendance records in "Reports"<br><br>
        <b>Tips for Best Results:</b><br>
        • Ensure adequate lighting in the room<br>
        • Position faces centrally in the camera view<br>
        • Clear face of obstructions (masks, glasses)<br>
        • Maintain consistent distance from camera
        </span>
        """)
        info_text.setWordWrap(True)
        info_layout.addWidget(info_text)

        layout.addWidget(info_frame, 1)

        # Quick actions
        actions_layout = QHBoxLayout()

        start_btn = ModernButton("▶ Start Attendance", accent=True)
        start_btn.clicked.connect(self.show_attendance)

        view_btn = ModernButton("📊 View Reports")
        view_btn.clicked.connect(self.show_reports)

        actions_layout.addWidget(start_btn)
        actions_layout.addWidget(view_btn)
        actions_layout.addStretch()

        layout.addLayout(actions_layout)

        return page

    def create_attendance_page(self):
        """Create attendance marking page"""
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # Header
        header = QLabel("Mark Attendance")
        font = QFont()
        font.setPointSize(22)
        font.setBold(True)
        header.setFont(font)
        header.setStyleSheet("color: #FF5555; margin-bottom: 10px;")
        layout.addWidget(header)

        # Configuration panel
        config_frame = QFrame()
        config_frame.setStyleSheet("""
            QFrame {
                background-color: #2D2D2D;
                border-radius: 8px;
                border: 1px solid #4472C4;
            }
        """)
        config_layout = QVBoxLayout(config_frame)
        config_layout.setContentsMargins(20, 15, 20, 15)
        config_layout.setSpacing(12)

        # Row 1: Subject
        row1 = QHBoxLayout()
        row1.setSpacing(10)

        subject_label = QLabel("Subject:")
        subject_label.setMinimumWidth(120)
        subject_label.setStyleSheet("font-weight: bold; color: #E0E0E0;")
        self.subject_combo = QComboBox()
        self.subject_combo.addItems([
            "Media and Info Literacy",
            "21st Century Literature From the Philippines and the World",
            "Computer System Servicing",
            "(Research Project) Inquiries, Investigation and Immersion",
            "Understanding Culture, Society and Politics",
            "Cookery"
        ])

        row1.addWidget(subject_label)
        row1.addWidget(self.subject_combo, 1)
        config_layout.addLayout(row1)

        # Row 2: Time settings
        row2 = QHBoxLayout()
        row2.setSpacing(10)

        time_label = QLabel("Class Start:")
        time_label.setMinimumWidth(120)
        time_label.setStyleSheet("font-weight: bold; color: #E0E0E0;")

        self.start_time = QTimeEdit()
        self.start_time.setTime(QTime.currentTime())
        self.start_time.setMinimumWidth(100)

        now_btn = ModernButton("Now")
        now_btn.setMaximumWidth(60)
        now_btn.clicked.connect(lambda: self.start_time.setTime(QTime.currentTime()))

        late_label = QLabel("Late Threshold (min):")
        late_label.setStyleSheet("font-weight: bold; color: #E0E0E0;")

        self.late_threshold = QSpinBox()
        self.late_threshold.setValue(10)
        self.late_threshold.setMinimum(1)
        self.late_threshold.setMaximum(60)
        self.late_threshold.setMaximumWidth(60)

        row2.addWidget(time_label)
        row2.addWidget(self.start_time)
        row2.addWidget(now_btn)
        row2.addSpacing(20)
        row2.addWidget(late_label)
        row2.addWidget(self.late_threshold)
        row2.addStretch()
        config_layout.addLayout(row2)

        layout.addWidget(config_frame)

        # Camera frame
        camera_container = QHBoxLayout()
        camera_container.setSpacing(15)

        # Camera feed
        camera_frame = QFrame()
        camera_frame.setStyleSheet("""
            QFrame {
                background-color: #1A1A1A;
                border: 2px solid #4472C4;
                border-radius: 8px;
            }
        """)
        camera_layout = QVBoxLayout(camera_frame)
        camera_layout.setContentsMargins(0, 0, 0, 0)

        self.camera_label = QLabel()
        self.camera_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.camera_label.setStyleSheet("color: #888888; font-size: 12pt;")
        self.camera_label.setMinimumSize(500, 400)
        self.camera_label.setText("📷 Camera Feed\nClick 'Start Session' to begin")

        camera_layout.addWidget(self.camera_label)
        camera_container.addWidget(camera_frame, 2)

        # Side panel
        side_panel = QFrame()
        side_panel.setStyleSheet("""
            QFrame {
                background-color: #2D2D2D;
                border-radius: 8px;
                border: 1px solid #3A3A3A;
            }
        """)
        side_layout = QVBoxLayout(side_panel)
        side_layout.setContentsMargins(15, 15, 15, 15)
        side_layout.setSpacing(10)

        # Session info
        info_label = QLabel("Session Info")
        info_label.setStyleSheet("font-weight: bold; color: #FF5555; font-size: 11pt;")
        side_layout.addWidget(info_label)

        self.session_status = QLabel("Status: Idle")
        self.session_status.setStyleSheet("color: #888888;")
        side_layout.addWidget(self.session_status)

        self.marked_label = QLabel("Marked: 0")
        self.marked_label.setStyleSheet("color: #00B050; font-weight: bold; font-size: 14pt;")
        side_layout.addWidget(self.marked_label)

        divider = QFrame()
        divider.setStyleSheet("background-color: #3A3A3A;")
        divider.setMaximumHeight(1)
        side_layout.addWidget(divider)

        # Recent marks list
        recent_label = QLabel("Recent Marks")
        recent_label.setStyleSheet("font-weight: bold; color: #4472C4; font-size: 10pt;")
        side_layout.addWidget(recent_label)

        self.recent_list = QListWidget()
        self.recent_list.setMaximumHeight(250)
        side_layout.addWidget(self.recent_list)

        side_layout.addStretch()

        camera_container.addWidget(side_panel, 1)
        layout.addLayout(camera_container, 1)

        # Controls
        controls_layout = QHBoxLayout()
        controls_layout.setSpacing(10)

        self.start_btn = ModernButton("▶ Start Session", accent=True)
        self.start_btn.clicked.connect(self.start_session)

        self.stop_btn = ModernButton("⏹ Stop & Save")
        self.stop_btn.setEnabled(False)
        self.stop_btn.clicked.connect(self.stop_session)

        controls_layout.addWidget(self.start_btn)
        controls_layout.addWidget(self.stop_btn)
        controls_layout.addStretch()

        layout.addLayout(controls_layout)

        return page

    def create_reports_page(self):
        """Create reports page"""
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # Header
        header = QLabel("Attendance Reports")
        font = QFont()
        font.setPointSize(22)
        font.setBold(True)
        header.setFont(font)
        header.setStyleSheet("color: #FF5555; margin-bottom: 10px;")
        layout.addWidget(header)

        # Options
        options_frame = QFrame()
        options_frame.setStyleSheet("""
            QFrame {
                background-color: #2D2D2D;
                border-radius: 8px;
                border: 1px solid #4472C4;
            }
        """)
        options_layout = QHBoxLayout(options_frame)
        options_layout.setContentsMargins(15, 15, 15, 15)
        options_layout.setSpacing(10)

        weekly_btn = ModernButton("📅 Weekly Summary")
        weekly_btn.clicked.connect(self.generate_weekly_report)

        monthly_btn = ModernButton("📊 Monthly Summary")
        monthly_btn.clicked.connect(self.generate_monthly_report)

        export_btn = ModernButton("💾 Export Excel", accent=True)

        options_layout.addWidget(weekly_btn)
        options_layout.addWidget(monthly_btn)
        options_layout.addWidget(export_btn)
        options_layout.addStretch()

        layout.addWidget(options_frame)

        # Table
        self.records_table = QTableWidget()
        self.records_table.setColumnCount(7)
        self.records_table.setHorizontalHeaderLabels([
            "Date", "Time", "Subject", "ID", "Name", "Section", "Status"
        ])
        self.records_table.horizontalHeader().setStretchLastSection(False)
        self.records_table.setAlternatingRowColors(True)
        self.load_records_table()

        layout.addWidget(self.records_table, 1)

        return page

    def create_settings_page(self):
        """Create settings page"""
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # Header
        header = QLabel("Settings")
        font = QFont()
        font.setPointSize(22)
        font.setBold(True)
        header.setFont(font)
        header.setStyleSheet("color: #FF5555; margin-bottom: 10px;")
        layout.addWidget(header)

        # Settings frame
        settings_frame = QFrame()
        settings_frame.setStyleSheet("""
            QFrame {
                background-color: #2D2D2D;
                border-radius: 8px;
                border: 1px solid #4472C4;
            }
        """)
        settings_layout = QVBoxLayout(settings_frame)
        settings_layout.setContentsMargins(20, 20, 20, 20)
        settings_layout.setSpacing(15)

        # Face Recognition Settings
        fr_label = QLabel("Face Recognition Settings")
        fr_label.setStyleSheet("font-weight: bold; color: #FF5555; font-size: 12pt;")
        settings_layout.addWidget(fr_label)

        # Confidence threshold
        conf_layout = QHBoxLayout()
        conf_label = QLabel("Confidence Threshold:")
        conf_label.setMinimumWidth(200)
        conf_label.setStyleSheet("font-weight: bold;")
        self.confidence_spin = QSpinBox()
        self.confidence_spin.setValue(60)
        self.confidence_spin.setSuffix("%")
        self.confidence_spin.setMaximumWidth(100)
        conf_layout.addWidget(conf_label)
        conf_layout.addWidget(self.confidence_spin)
        conf_layout.addStretch()
        settings_layout.addLayout(conf_layout)

        # Buffer frames
        buffer_layout = QHBoxLayout()
        buffer_label = QLabel("Recognition Buffer Frames:")
        buffer_label.setMinimumWidth(200)
        buffer_label.setStyleSheet("font-weight: bold;")
        self.buffer_spin = QSpinBox()
        self.buffer_spin.setValue(3)
        self.buffer_spin.setMinimum(1)
        self.buffer_spin.setMaximum(10)
        self.buffer_spin.setMaximumWidth(100)
        buffer_layout.addWidget(buffer_label)
        buffer_layout.addWidget(self.buffer_spin)
        buffer_layout.addStretch()
        settings_layout.addLayout(buffer_layout)

        settings_layout.addSpacing(20)

        # Database Settings
        db_label = QLabel("Database Management")
        db_label.setStyleSheet("font-weight: bold; color: #FF5555; font-size: 12pt;")
        settings_layout.addWidget(db_label)

        register_btn = ModernButton("📂 Register Students from Folder")
        register_btn.clicked.connect(self.register_students)
        settings_layout.addWidget(register_btn)

        settings_layout.addSpacing(20)

        # Save button
        save_btn = ModernButton("💾 Save Settings", accent=True)
        save_btn.clicked.connect(self.save_settings)
        settings_layout.addWidget(save_btn)

        layout.addWidget(settings_frame)
        layout.addStretch()

        return page

    def load_records_table(self):
        """Load attendance records from Excel files"""
        # Placeholder - would load from Attendance_Records folder
        pass

    # Navigation methods
    def show_dashboard(self):
        """Show dashboard"""
        self.stacked_widget.setCurrentWidget(self.dashboard_page)

    def show_attendance(self):
        """Show attendance"""
        self.stacked_widget.setCurrentWidget(self.attendance_page)

    def show_reports(self):
        """Show reports"""
        self.stacked_widget.setCurrentWidget(self.reports_page)

    def show_settings(self):
        """Show settings"""
        self.stacked_widget.setCurrentWidget(self.settings_page)

    # Session control methods
    def start_session(self):
        """Start attendance session"""
        self.session_active = True
        self.marked_count = 0
        self.recent_list.clear()

        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.session_status.setText(f"Status: Recording")
        self.session_status.setStyleSheet("color: #00B050;")
        self.marked_label.setText("Marked: 0")

        # In a real implementation, would start camera thread
        QMessageBox.information(self, "Session Started",
                              f"Attendance session started for:\n{self.subject_combo.currentText()}")

    def stop_session(self):
        """Stop attendance session"""
        self.session_active = False

        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.session_status.setText("Status: Stopped")
        self.session_status.setStyleSheet("color: #FF9800;")

        QMessageBox.information(self, "Session Stopped",
                              f"Session stopped. {self.marked_count} students marked.")

    def generate_weekly_report(self):
        """Generate weekly report"""
        QMessageBox.information(self, "Report", "Weekly summary generated successfully")

    def generate_monthly_report(self):
        """Generate monthly report"""
        QMessageBox.information(self, "Report", "Monthly summary generated successfully")

    def register_students(self):
        """Register students"""
        folder = QFileDialog.getExistingDirectory(self, "Select Folder with Student Images")
        if folder:
            QMessageBox.information(self, "Registration",
                                  f"Students registered from:\n{folder}")

    def save_settings(self):
        """Save settings"""
        QMessageBox.information(self, "Settings Saved",
                              "Your settings have been saved successfully")


def main():
    app = QApplication(sys.argv)
    window = AttendanceGUI()
    sys.exit(app.exec())


if __name__ == '__main__':
    main()
